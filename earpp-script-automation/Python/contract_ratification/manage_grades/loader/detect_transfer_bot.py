import openpyxl
from loader.config import Colors, base_loaders, column_names_coordinates, column_mapping_dict
import os
import configparser
from pathlib import Path
environment = os.environ.get('ENV', 'DEV')
import xlwings as xw
import pandas as pd

class TransferExcelData:
    def __init__(self, psbuild_type, file_id):
        config = configparser.ConfigParser()
        config_file_path = os.path.join(str(Path(__file__).parents[1]),'manage_grades.ini')
        config.read(config_file_path)
        
        self.psbuild_type = psbuild_type
        
        root_folder = str(Path(__file__).parents[1])
        process_folder = str(config.get(f"{environment}.ManageGrades", 'processfolder' + psbuild_type))
        
        self.path_excel = f'{process_folder}/Loader_{file_id}.xlsx'
        self.cloud_loader_path = f"{root_folder}/{base_loaders[psbuild_type]}"
        self.cloud_build_path = f'{process_folder}/CBuild_{file_id}.xlsx'
        
        # self.wb = openpyxl.load_workbook(self.path_excel)
        # self.sheet = self.wb.active
        table = str.maketrans("[]", "()")
        book_key = f'Loader_{file_id}.xlsx'.translate(table)
        self.wb = wb = xw.apps[xw.apps.keys()[0]].books[book_key]
        self.sheet = wb.sheets.active
    
    # getting cell range base on first and last column of cloud build.
    def get_cell_coordinate(self, text_value):
            for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=100, max_col=100):
                for cell in row:
                    if cell.value == text_value:
                        return cell.coordinate
                    
    #getting values and adding it on a list                
    def cell_values_between(self, range_left, range_right):
        list_val = []
        row_focus = self.sheet[range_left:range_right]
        for cell in row_focus:
            for x in cell:
                list_val.append(x.value)

        return list_val

    #getting all range and cell values.
    def range_and_values(self, start_value, end_value):
        min = self.get_cell_coordinate(start_value)
        print(f"Leftmost cell = {Colors.YELLOW}{min}{Colors.RESET}")
        max = self.get_cell_coordinate(end_value)
        print(f"Rightmost cell = {Colors.YELLOW}{max}{Colors.RESET}")
        listing = self.cell_values_between(min,max)
        print(f"{Colors.BLUE}{listing}{Colors.RESET}")
        return listing
    
    #getting number of columns
    def number_of_columns(self, loader_header_list, cloud_loader_header_list):
        print(f"LOADER: Focused header columns = {Colors.BLUE}{len(loader_header_list)}{Colors.RESET}")
        print(f"CLOUD BUILD: Focused header columns = {Colors.BLUE}{len(cloud_loader_header_list)}{Colors.RESET}")
        if len(loader_header_list) == len(cloud_loader_header_list):
            print(f"{Colors.GREEN}PASSED: No Changes In The Number of Columns{Colors.RESET}")
        else:
            print(f"{Colors.RED}FAILED: Need to Update Number of Columns{Colors.RESET}")
    
    #exact match on every cell value
    def exact_match(self, loader_header_list, cloud_loader_header_list):
        length_checker = []
        for i in loader_header_list:
            for j in cloud_loader_header_list: 
                if i == j:
                    print(f"{Colors.GREEN}MATCHED:{Colors.RESET} {i} {Colors.GREEN}={Colors.RESET} {j}")
                    length_checker.append(i)
                

        if len(length_checker) == len(loader_header_list):
            print(f"{Colors.GREEN}{len(length_checker)}/{len(cloud_loader_header_list)} Matches{Colors.RESET}")
            print((f"{Colors.GREEN}PASSED: No Changes In The Loader File{Colors.RESET}"))
            return True
        else:
            print(f"{Colors.RED}{len(length_checker)}/{len(cloud_loader_header_list)} Matches{Colors.RESET}")
            print(f"{Colors.RED}Loader File Columns Doesn't Matched Cloud Build {Colors.RESET}")
            print(f"{Colors.RED}FAILED: Need to update MyACI Loader file {Colors.RESET}")
            return False

    def validate_loader_file(self):
        print("Validating Loader File Columns")
        column_names_range = column_names_coordinates[self.psbuild_type]
        new_loader_column_names = self.sheet[column_names_range].value
    
        base_loader_sheet = openpyxl.load_workbook(self.cloud_loader_path, read_only=True).active
        base_loader_column_names = [col_name.value for col_name in list(base_loader_sheet[column_names_range][0])]
        
        
        print(f"{Colors.GREEN}{len(new_loader_column_names)}/{len(base_loader_column_names)} Matches{Colors.RESET}")
        
        if base_loader_column_names == new_loader_column_names:
            print((f"{Colors.GREEN}PASSED: No Changes In The Loader File{Colors.RESET}"))
            return True
        
        print(f"{Colors.RED}Loader File Columns Doesn't Matched Cloud Build {Colors.RESET}")
        print(f"{Colors.RED}FAILED: Need to update MyACI Loader file {Colors.RESET}")
        return False
        #looping the header columns and rows under each iteration inside Cloud Build.

    def loader_conditional(self):
        if self.validate_loader_file() == True:
            self.copy_cb_to_loader()  
        else:
            print(f"{Colors.RED}Loader mismatched, unable to copy cloud build.{Colors.RESET}")
            print(f"{Colors.RED}About to exit program.{Colors.RESET}")


    def copy_cb_to_loader(self):
        print("Copying Cloud Build data to Loader file")
        column_names_range = column_names_coordinates[self.psbuild_type]
        loader_col_names = self.sheet[column_names_range]
        
        cbuild_df = pd.read_excel(self.cloud_build_path, dtype=str)
        cbuild_col_names = cbuild_df.columns.to_list()
        print(cbuild_col_names)

        print(loader_col_names.value)
        matched_columns_name = [column_mapping_dict[self.psbuild_type][loader_col_name.value] for loader_col_name in loader_col_names]
        cbuild_df = cbuild_df.reindex(matched_columns_name, axis=1)
        print(cbuild_df.columns.values)

        cbuild_df_len = len(cbuild_df.index)
        first_data_row_cell = (loader_col_names.row + 1, loader_col_names.column)
        last_data_row_cell = (loader_col_names.row + 1 + cbuild_df_len, loader_col_names.last_cell.column)
        print(first_data_row_cell, last_data_row_cell)
        print(f"{loader_col_names.row + 1}:{loader_col_names.row + 1 + cbuild_df_len}")
        # self.sheet.range(f"{loader_col_names.row + 1}:{loader_col_names.row + cbuild_df_len}").insert(shift='down')
        # self.sheet.range(f"{loader_col_names.row + cbuild_df_len + 1}:{loader_col_names.row + cbuild_df_len + 2}").delete()
        # self.sheet.range(first_data_row_cell, last_data_row_cell).value = [['' for x in range(cbuild_df.shape[1])] for x in range(cbuild_df.shape[0])]
        # self.sheet.range(first_data_row_cell, last_data_row_cell).clear_contents()

        try:
            self.sheet.range(first_data_row_cell, last_data_row_cell).value = cbuild_df.values.tolist()
        except Exception as e:
            self.sheet.range(first_data_row_cell, last_data_row_cell).value = cbuild_df.values.tolist()
        # data_sheet_range = self.sheet[]
        # print(data_sheet_range.last_cell.row + cbuild_df_len, data_sheet_range.last_cell.column)

        # for loader_col_name in loader_col_names:
        #     print(loader_col_name.value)
        #     match_column = column_mapping_dict[self.psbuild_type][loader_col_name.value]

        #     cbuild_col = cbuild_df[match_column]

        #     if 'date' in match_column.lower():
        #         cbuild_col = pd.to_datetime(cbuild_col).dt.strftime(date_format="%m/%d/%Y")
        #     cbuild_col = cbuild_col.fillna('')

        #     for index, value in cbuild_col.items():
        #         print(loader_col_name.row + index, loader_col_name.column -1, value)
        #         self.sheet[loader_col_name.row + index, loader_col_name.column - 1].value = str(value)

    #ariel start
    def find_target_value_position_in_column(self, column_letter, search_string):
               
        last_row = self.sheet.range(column_letter + str(self.sheet.cells.last_cell.row)).end('up').row

        #Loop through the column to find the matching string
        for row in range(1, last_row + 1):
            cell_value = self.sheet.cells(row, column_letter).value
            if cell_value == search_string:
                return row + 1            

   
    def are_cells_in_column_contain_this_str(self, column_letter, start_row, str_to_search):

        last_row = self.sheet.range(column_letter + str(self.sheet.cells.last_cell.row)).end('up').row

        for row in range(start_row, last_row + 1):
            cell_value = self.sheet.cells(row, column_letter).value
            if cell_value == str_to_search:
                return True
            
        return False

    def status_checker(self, loader_automator):
        column_letter = 'D'     # Upload Progress Column

        str_to_search = 'Upload Progress'

        start_index = self.find_target_value_position_in_column(column_letter, str_to_search)

        while self.are_cells_in_column_contain_this_str(column_letter, start_index, 'In Progress') or self.are_cells_in_column_contain_this_str(column_letter, start_index, 'Ready to Process'):
            loader_automator.refresh_upload()

            print('Waiting for Upload Progress to complete...')

        print("Succesfully Uploaded")

            #existing_atb_unprotect.excel_upload(pass excel title)
    #ariel end

                    
if __name__=="__main__":
    pass
    

    


