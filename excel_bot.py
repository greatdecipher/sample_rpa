import openpyxl
from config import ExcelFilePath, Colors, CellValues, Labels

class OpenExcel:
    def __init__(self, path_excel):
        self.loader_myaci = None
        self.path_excel = path_excel
        self.wb = openpyxl.load_workbook(self.path_excel)
        self.sheet = self.wb.active
    
    # getting cell range base on first and last column of cloud build.
    def get_cell_coordinate(self, text_value):
            for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=100, max_col=100):
                for cell in row:
                    if cell.value == text_value:
                        return cell.coordinate
                    
    #getting values and adding it on a list                
    def cell_values_between(self, range_left, range_right):
        list_val = []
        row_focus = self.sheet[range_left:range_right]
        for cell in row_focus:
            for x in cell:
                list_val.append(x.value)

        return list_val

    #getting all range and cell values.
    def range_and_values(self, start_value, end_value):
        min = self.get_cell_coordinate(start_value)
        print(f"Leftmost cell = {Colors.YELLOW}{min}{Colors.RESET}")
        max = self.get_cell_coordinate(end_value)
        print(f"Rightmost cell = {Colors.YELLOW}{max}{Colors.RESET}")
        listing = self.cell_values_between(min,max)
        print(f"{Colors.BLUE}{listing}{Colors.RESET}")
        return listing
    
    #getting number of columns
    def number_of_columns(self, loader_list, cloud_loader_list):
        print(f"LOADER: Focused header columns = {Colors.BLUE}{len(loader_list)}{Colors.RESET}")
        print(f"CLOUD BUILD: Focused header columns = {Colors.BLUE}{len(cloud_loader_list)}{Colors.RESET}")
        if len(loader_list) == len(cloud_loader_list):
            print(f"{Colors.GREEN}PASSED: No Changes In The Number of Columns{Colors.RESET}")
        else:
            print(f"{Colors.RED}FAILED: Need to Update Number of Columns{Colors.RESET}")
    
    #partial match on every cell value
    def exact_match(self, loader_list, cloud_loader_list):
        length_checker = []
        for i in loader_list:
            for j in cloud_loader_list: 
                if i == j:
                    print(f"{Colors.GREEN}MATCHED:{Colors.RESET} {i} {Colors.GREEN}={Colors.RESET} {j}")
                    length_checker.append(i)
                

        if len(length_checker) == len(loader_list):
            print(f"{Colors.GREEN}{len(length_checker)}/{len(cloud_loader_list)} Matches{Colors.RESET}")
            print((f"{Colors.GREEN}PASSED: No Changes In The Loader File{Colors.RESET}"))
        else:
            print(f"{Colors.RED}{len(length_checker)}/{len(cloud_loader_list)} Matches{Colors.RESET}")
            print(f"{Colors.RED}Loader File Columns Doesn't Matched Cloud Build {Colors.RESET}")
            print(f"{Colors.RED}FAILED: Need to update MyACI Loader file {Colors.RESET}")

    def validate_loader_file(self):
        loader_list = self.range_and_values(CellValues.leftmost_col,
                        CellValues.rightmost_col)   
        print('='* 100)
        #based loader file
        match self.path_excel:
            case ExcelFilePath.test_existing_atb_loader:
                self.cloud_loader = OpenExcel(ExcelFilePath.test_existing_cloud)
                self.cloud_build = OpenExcel(ExcelFilePath.cloud_build_existing)
            #match file to test
            case ExcelFilePath.match_existing_loader:
                self.cloud_loader = OpenExcel(ExcelFilePath.test_existing_cloud)
                self.cloud_build = OpenExcel(ExcelFilePath.cloud_build_existing)
            case ExcelFilePath.test_with_atb_loader:
                self.cloud_loader = OpenExcel(ExcelFilePath.test_with_cloud)
                self.cloud_build = OpenExcel(ExcelFilePath.cloud_build_with)
            case ExcelFilePath.test_without_atb_loader:
                self.cloud_loader = OpenExcel(ExcelFilePath.test_without_cloud)
                self.cloud_build = OpenExcel(ExcelFilePath.cloud_build_without)

        cloud_loader_list = self.cloud_loader.range_and_values(CellValues.leftmost_col,
                        CellValues.rightmost_col)   
        
        #number of columns checker
        self.number_of_columns(loader_list,cloud_loader_list)

        print('='* 100)
        #exact match checker for loader and cloud loader 'bound to fail'
        self.exact_match(loader_list, cloud_loader_list)


        #looping the header columns and rows under each iteration inside Cloud Build.
    def copy_cb_to_loader(self):
        self.coordinate_list = []
        cloud_focused_headers = self.cloud_build.range_and_values("Effective Start Date (Based from report Effective Start Date) mm/dd/yyyy",
                        "ATB Rate (New ATB Rate)")
        for i in cloud_focused_headers:
            coordinate = self.cloud_build.get_cell_coordinate(i)
            self.coordinate_list.append(coordinate)
        print(self.coordinate_list)
        

        for i in self.coordinate_list:
            for cell in self.cloud_build.sheet.iter_rows(min_row = 2,max_col = 20, max_row=100):
                for j in cell:
                    print('Printing from ' + 'Column:' + str(j.column) +', Row:'+ str(j.row))
                    self.sheet.cell(j.row + 25, column=j.column + 5, value=j.value)

        self.wb.save('./test_files\sample30.xlsx')
                    
if __name__=="__main__":
    existing_atb = OpenExcel(ExcelFilePath.test_existing_atb_loader)
    with_atb = OpenExcel(ExcelFilePath.test_with_atb_loader)
    without_atb = OpenExcel(ExcelFilePath.test_without_atb_loader)
    matched_existing = OpenExcel(ExcelFilePath.match_existing_loader)

    #without_atb.validate_loader_file()
    #with_atb.validate_loader_file()
    #existing_atb.validate_loader_file()
    matched_existing.validate_loader_file()
    #matched_existing.copy_cb_to_loader()
    #matched_existing.copy_cloud_build()
    
    


